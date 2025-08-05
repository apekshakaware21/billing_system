import openpyxl
import os


DATA_FILE = 'customer_records.xlsx'
BILL_FILE = 'bill_book.xlsx'
CGST_RATE = 0.09
SGST_RATE = 0.09

def initialize_excel():
    if not os.path.exists(DATA_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["Customer ID", "Company Name", "Phone No", "Email ID"])
        wb.save(DATA_FILE)

    if not os.path.exists(BILL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bills"
        ws.append(["Customer ID", "Purchased Items", "Base Amount", "CGST", "SGST", "Total Payable"])
        wb.save(BILL_FILE)

def admin_panel():
    while True:
        print("\n===== Admin Panel =====")
        print("1. Add Customer Record")
        print("2. View All Records")
        print("3. Search Customer by ID")
        print("4. Generate Bill")
        print("5. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            add_customer()
        elif choice == '2':
            view_all_customers()
        elif choice == '3':
            search_customer()
        elif choice == '4':
            generate_bill()
        elif choice == '5':
            print("Exiting Admin Panel.")
            break
        else:
            print("Invalid choice. Try again.")

def add_customer():
    cid = input("Enter Customer ID: ")
    name = input("Enter Company Name: ")
    phone = input("Enter Phone Number: ")
    email = input("Enter Email ID: ")

    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    ws.append([cid, name, phone, email])
    wb.save(DATA_FILE)
    print("Customer record added successfully!")

def view_all_customers():
    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    print("\nCustomer Records:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)

def search_customer():
    cid = input("Enter Customer ID to search: ")
    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == cid:
            print("Record Found:", row)
            found = True
            break
    if not found:
        print("Customer ID not found.")

def generate_bill():
    cid = input("Enter Customer ID: ")
    items = input("Enter Purchased Items: ")
    try:
        base_amount = float(input("Enter Base Amount: "))
    except ValueError:
        print("Invalid amount entered.")
        return

    cgst = base_amount * CGST_RATE
    sgst = base_amount * SGST_RATE
    total = base_amount + cgst + sgst

    wb = openpyxl.load_workbook(BILL_FILE)
    ws = wb.active
    ws.append([cid, items, base_amount, cgst, sgst, total])
    wb.save(BILL_FILE)
    print("Bill generated and saved successfully!")

if __name__ == "__main__":
    initialize_excel()
    admin_panel()
